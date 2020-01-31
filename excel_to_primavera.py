import openpyxl
import configparser
import main as m


def user_id_from_name(serv, usr, passw, name):
    name_first = name.split(' ')[0]
    name_last = name.split(' ')[-1]
    request_data = {'Field': ['ObjectId', 'Name']
        ,'Filter': "PersonalName like '%%%s%%'" % name_first
        ,'Filter': "PersonalName like '%%%s%%'" % name_last
    }
    if name:  # if it's not None
        synched = m.soap_request(request_data, serv, 'UserService', 'ReadUsers', usr, passw)
        return synched[0]['ObjectId'], synched[0]['Name']
    else:
        return None, None


def post_note(serv, usr, passw, parent, code, note):
    request_data = {'ActivityNote': {'Note': note,
                                     'ActivityObjectId': parent,
                                     'NotebookTopicObjectId': code  # 43 and 38 this better never change
                                     }}
    synched = m.soap_request(request_data, serv, 'ActivityNoteService', 'CreateActivityNotes', usr, passw)
    return synched

# read config and get primavera login info
parser = configparser.ConfigParser()
with open('login') as configfile:
    parser.read_file(configfile)
primadict=parser['primavera-section']
primauser=primadict['user']
primapasswd=primadict['passwd']
primaserver=primadict['server']
# read tool config
tool_dict = parser['tool-settings']
tool_log = tool_dict['loglevel']
tool_fixer = tool_dict['fix']
if tool_fixer.lower() in ['true', '1', 't', 'y', 'yes', 'yeah', 'yup', 'certainly', 'uh-huh']:
    tool_fixer = True
else:
    tool_fixer = False


if __name__ == '__main__':
    # preload file and worksheet
    workbook = openpyxl.load_workbook('input.xlsx', data_only=True)
    acts_sheet = workbook.get_sheet_by_name('Epic Suggestions')
    steps_sheet = workbook.get_sheet_by_name('TaskStory Suggestion')

    # go through activity entries in the file
    # because max_row method counts rows even after they've been wiped, it's time to go back to the old VBA ways...
    i = 2
    while True:
        import_check = acts_sheet.cell(row=i, column=1).value
        if import_check == 'TRUE' or import_check == True:
            continue  # skip TRUE items
        if import_check == '' or import_check == None or import_check == ' ':
            break  # stop execution
        act_name = acts_sheet.cell(row=i, column=2).value
        owner = acts_sheet.cell(row=i, column=3).value
        api_owner_obj, api_owner_name = user_id_from_name(primaserver, primauser, primapasswd, owner)
        purpose = acts_sheet.cell(row=i, column=6).value
        scope = acts_sheet.cell(row=i, column=7).value
        # Activity creation code
        request_data = {'Activity': {'Name': act_name,
                                     'AtCompletionDuration': 0,
                                     'CalendarObjectId': 638,  # NCSA Standard with Holidays, ok to hardcode
                                     'ProjectObjectId': m.actual_baseline(primaserver, primauser, primapasswd),
                                     # 408 is the one currently active
                                     'ProjectId': 'LSST MREFC',
                                     'WBSObjectId': 4597,
                                     'WBSPath': 'LSST MREFC.MREFC.LSST Construction.Test WBS',  # the big dump
                                     'OwnerIDArray': api_owner_obj,
                                     'OwnerNamesArray': api_owner_name
                                     }}
        synched = m.soap_request(request_data, primaserver, 'ActivityService', 'CreateActivities', primauser, primapasswd)
        created_activity = synched[0]
        print(created_activity)
        # post descriptions to the new activity
        # purpose
        synched = post_note(primaserver, primauser, primapasswd, created_activity, 38, purpose)
        # scope
        synched = post_note(primaserver, primauser, primapasswd, created_activity, 43, scope)
        # go through all steps and find relevant ones
        r = 2
        while True:
            step_name = steps_sheet.cell(row=r, column=2).value
            if step_name == '' or step_name == None or step_name == ' ':
                break
            if steps_sheet.cell(row=r, column=1).value == act_name:
                step_desc = steps_sheet.cell(row=r, column=7).value
                try:
                    step_pts = int(steps_sheet.cell(row=r, column=5).value)/4
                except TypeError:
                    step_pts = 0
                # Step creation code
                request_data = {'ActivityStep': {'ActivityObjectId': created_activity,
                                             'Description': step_desc,
                                             'Name': step_name,
                                             'Weight': step_pts}}
                synched = m.soap_request(request_data, primaserver, 'ActivityStepService', 'CreateActivitySteps', primauser, primapasswd)
                # created_activity = synched[0]
            r += 1
        i += 1

    # 41186
